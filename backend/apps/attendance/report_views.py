"""Report generation and export views for attendance data."""
import io
from datetime import timedelta

from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.models import Employee
from apps.core.permissions import IsTenantAdmin, IsTenantMember

from .models import AttendanceRecord


class _ReportMixin:
    """Shared logic for building attendance report querysets."""

    @staticmethod
    def _get_filters(request, tenant):
        date_from = request.query_params.get(
            "date_from", str(timezone.localdate() - timedelta(days=30))
        )
        date_to = request.query_params.get("date_to", str(timezone.localdate()))
        status_filter = request.query_params.get("status")
        department_id = request.query_params.get("department")
        employee_id = request.query_params.get("employee_id")
        return date_from, date_to, status_filter, department_id, employee_id

    @staticmethod
    def _build_queryset(tenant, date_from, date_to, status_filter, department_id, employee_id):
        qs = (
            AttendanceRecord.objects.filter(
                tenant=tenant, date__gte=date_from, date__lte=date_to, is_deleted=False
            )
            .select_related("employee__user", "employee__department", "shift")
            .order_by("date", "clock_in_time")
        )
        if status_filter:
            qs = qs.filter(status=status_filter)
        if department_id:
            qs = qs.filter(employee__department_id=department_id)
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs

    @staticmethod
    def _build_summary(tenant, date_from, date_to, status_filter, department_id, employee_id):
        qs = AttendanceRecord.objects.filter(
            tenant=tenant, date__gte=date_from, date__lte=date_to, is_deleted=False
        )
        if status_filter:
            qs = qs.filter(status=status_filter)
        if department_id:
            qs = qs.filter(employee__department_id=department_id)
        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        employees = (
            qs.values(
                "employee__id",
                "employee__employee_id",
                "employee__user__first_name",
                "employee__user__last_name",
                "employee__department__name",
            )
            .annotate(
                total_days=Count("id"),
                present_count=Count("id", filter=Q(status="PRESENT")),
                absent_count=Count("id", filter=Q(status="ABSENT")),
                late_count=Count("id", filter=Q(status="LATE")),
                early_departure_count=Count("id", filter=Q(status="EARLY_DEPARTURE")),
                half_day_count=Count("id", filter=Q(status="HALF_DAY")),
                leave_count=Count("id", filter=Q(status="ON_LEAVE")),
            )
            .order_by("employee__user__first_name")
        )

        summaries = []
        for emp in employees:
            emp_records = qs.filter(employee__id=emp["employee__id"], clock_out_time__isnull=False)
            total_hours = sum(r.duration_hours or 0 for r in emp_records)
            days = emp["total_days"] or 1
            summaries.append(
                {
                    "employee_id": emp["employee__employee_id"] or str(emp["employee__id"])[:8],
                    "employee_name": f"{emp['employee__user__first_name']} {emp['employee__user__last_name']}",
                    "department": emp["employee__department__name"] or "—",
                    "total_days": emp["total_days"],
                    "present": emp["present_count"],
                    "absent": emp["absent_count"],
                    "late": emp["late_count"],
                    "early_departure": emp["early_departure_count"],
                    "half_day": emp["half_day_count"],
                    "on_leave": emp["leave_count"],
                    "total_hours": round(total_hours, 2),
                    "avg_hours": round(total_hours / days, 2),
                }
            )
        return summaries


# -------------------------------------------------------------------
# JSON report data (used by frontend table)
# -------------------------------------------------------------------
class AttendanceReportView(_ReportMixin, APIView):
    """Return attendance report data as JSON with filters."""

    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request, tenant_slug):
        tenant = request.tenant
        date_from, date_to, status_filter, department_id, employee_id = self._get_filters(
            request, tenant
        )
        report_type = request.query_params.get("type", "summary")

        if report_type == "detailed":
            qs = self._build_queryset(
                tenant, date_from, date_to, status_filter, department_id, employee_id
            )
            records = []
            for r in qs[:500]:
                records.append(
                    {
                        "id": str(r.id),
                        "date": str(r.date),
                        "employee_name": r.employee.user.full_name if r.employee and r.employee.user else "—",
                        "employee_id": r.employee.employee_id or "",
                        "department": r.employee.department.name if r.employee and r.employee.department else "—",
                        "shift": r.shift.name if r.shift else "—",
                        "clock_in": r.clock_in_time.strftime("%H:%M") if r.clock_in_time else "—",
                        "clock_out": r.clock_out_time.strftime("%H:%M") if r.clock_out_time else "—",
                        "status": r.status,
                        "duration": f"{r.duration_hours:.1f}h" if r.duration_hours else "—",
                        "geofence_valid": r.geofence_valid,
                    }
                )
            return Response({"success": True, "data": records})

        # Default: summary
        summaries = self._build_summary(
            tenant, date_from, date_to, status_filter, department_id, employee_id
        )
        return Response({"success": True, "data": summaries})


# -------------------------------------------------------------------
# Excel Export
# -------------------------------------------------------------------
class ExportExcelView(_ReportMixin, APIView):
    """Export attendance report as an Excel (.xlsx) file."""

    permission_classes = [IsAuthenticated, IsTenantAdmin]

    def get(self, request, tenant_slug):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        tenant = request.tenant
        date_from, date_to, status_filter, department_id, employee_id = self._get_filters(
            request, tenant
        )
        report_type = request.query_params.get("type", "summary")

        wb = Workbook()
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        if report_type == "detailed":
            ws.title = "Attendance Details"
            headers = [
                "Date", "Employee ID", "Employee Name", "Department",
                "Shift", "Clock In", "Clock Out", "Status", "Duration",
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            qs = self._build_queryset(
                tenant, date_from, date_to, status_filter, department_id, employee_id
            )
            for r in qs:
                ws.append([
                    str(r.date),
                    r.employee.employee_id or "",
                    r.employee.user.full_name if r.employee and r.employee.user else "",
                    r.employee.department.name if r.employee and r.employee.department else "",
                    r.shift.name if r.shift else "",
                    r.clock_in_time.strftime("%H:%M") if r.clock_in_time else "",
                    r.clock_out_time.strftime("%H:%M") if r.clock_out_time else "",
                    r.status,
                    f"{r.duration_hours:.1f}" if r.duration_hours else "",
                ])
        else:
            ws.title = "Attendance Summary"
            headers = [
                "Employee ID", "Employee Name", "Department", "Total Days",
                "Present", "Absent", "Late", "Early Dep.", "Half Day",
                "On Leave", "Total Hours", "Avg Hours/Day",
            ]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            summaries = self._build_summary(
                tenant, date_from, date_to, status_filter, department_id, employee_id
            )
            for s in summaries:
                ws.append([
                    s["employee_id"], s["employee_name"], s["department"],
                    s["total_days"], s["present"], s["absent"], s["late"],
                    s["early_departure"], s["half_day"], s["on_leave"],
                    s["total_hours"], s["avg_hours"],
                ])

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"attendance_report_{date_from}_to_{date_to}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# -------------------------------------------------------------------
# PDF Export
# -------------------------------------------------------------------
class ExportPdfView(_ReportMixin, APIView):
    """Export attendance report as a PDF file."""

    permission_classes = [IsAuthenticated, IsTenantAdmin]

    def get(self, request, tenant_slug):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        tenant = request.tenant
        date_from, date_to, status_filter, department_id, employee_id = self._get_filters(
            request, tenant
        )
        report_type = request.query_params.get("type", "summary")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_text = f"Attendance Report — {tenant.name}"
        elements.append(Paragraph(title_text, styles["Title"]))
        elements.append(Paragraph(f"Period: {date_from} to {date_to}", styles["Normal"]))
        elements.append(Spacer(1, 10 * mm))

        if report_type == "detailed":
            headers = ["Date", "Emp ID", "Name", "Dept", "Shift", "In", "Out", "Status", "Hours"]
            qs = self._build_queryset(
                tenant, date_from, date_to, status_filter, department_id, employee_id
            )
            data = [headers]
            for r in qs[:1000]:
                data.append([
                    str(r.date),
                    (r.employee.employee_id or "")[:10],
                    (r.employee.user.full_name if r.employee and r.employee.user else "")[:20],
                    (r.employee.department.name if r.employee and r.employee.department else "")[:15],
                    (r.shift.name if r.shift else "")[:10],
                    r.clock_in_time.strftime("%H:%M") if r.clock_in_time else "—",
                    r.clock_out_time.strftime("%H:%M") if r.clock_out_time else "—",
                    r.status[:10],
                    f"{r.duration_hours:.1f}" if r.duration_hours else "—",
                ])
        else:
            headers = [
                "Emp ID", "Name", "Dept", "Days", "Present",
                "Absent", "Late", "Early", "Leave", "Hours", "Avg",
            ]
            summaries = self._build_summary(
                tenant, date_from, date_to, status_filter, department_id, employee_id
            )
            data = [headers]
            for s in summaries:
                data.append([
                    str(s["employee_id"])[:10],
                    s["employee_name"][:20],
                    s["department"][:15],
                    s["total_days"],
                    s["present"],
                    s["absent"],
                    s["late"],
                    s["early_departure"],
                    s["on_leave"],
                    s["total_hours"],
                    s["avg_hours"],
                ])

        if len(data) > 1:
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(table)
        else:
            elements.append(Paragraph("No data found for the selected filters.", styles["Normal"]))

        doc.build(elements)
        buf.seek(0)

        filename = f"attendance_report_{date_from}_to_{date_to}.pdf"
        response = HttpResponse(buf.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
